"""
*******IMPORTANT*******
The purpose of this file is to collect and collate data from the various datasets, process it into
a useful form, then finally store it in the cleaned_data.csv file. TO DO THIS, IT PINGS A GEOLOCATOR
API WITH COORDINATES USED IN DATA GATHERING. THIS PROCESS WILL TAKE OVER FIVE (5) HOURS. UNLESS YOU
ARE OK WITH THAT, DO NOT RUN THIS FILE.

Data set notes:
- ds.xml is the DineSafe dataset, detailing all restaurants in the city by address and infractions
- ICES-COVID19-Vaccination-Data-by-FSA.xlsx is a spreadsheet detailing COVID-19 cases,
hospitalizations and deaths per population, as well as what percent of people have recieved at least
1, 2 and 3 doses of vaccine. Some data values are suppressed to prevent identifiability.
- T120120211212055123.csv details the population of each FSA.
"""
from geopy.geocoders import Nominatim  # Library for pinging Nominatim geolocator API
import openpyxl  # Library for reading data from excel spreadsheet
import csv  # Library for reading and writing to .csv files
from xml.dom import minidom  # Library for reading .xml files
import unidecode  # Library for getting rid of weird unicode characters in the aformentioned
# .xml files that were breaking my reader
from geopy.extra.rate_limiter import RateLimiter  # Library for ratelimiting my pings to the
# Nominatim geolocator API, which allows a maximum of one request per second.
from FSA import FSA  # My FSA data class, which allows me to create FSA (forward sortation area
# objects I use to store data)

# data dict maps FSA names to FSA data types. Easiest way to store and access info while the data
# is processed for all 96 Toronto FSAs.
data = {}

# First data batch: the Institude for Clinical Evaluative Studies (an Ontario health data research
# centre)'s COVID-19 vaccination data spreadsheet.
path = 'ICES-COVID19-Vaccination-Data-by-FSA.xlsx'
ICES_wb_obj = openpyxl.load_workbook(path)
ICES_1_dose_sheet_obj = ICES_wb_obj['At least 1 Dose  by FSA']


sheet_row = 272  # First row containing Toronto-relevant data
while sheet_row <= 367:  # Passing through all Toronto postal codes in the data set.
    FSA_cell = ICES_1_dose_sheet_obj.cell(row=sheet_row, column=1)
    cases_per_100_cell = ICES_1_dose_sheet_obj.cell(row=sheet_row, column=3)
    hospitalizations_per_1000_cell = ICES_1_dose_sheet_obj.cell(row=sheet_row, column=4)
    deaths_per_1000_cell = ICES_1_dose_sheet_obj.cell(row=sheet_row, column=5)
    percent_1_dose_cell = ICES_1_dose_sheet_obj.cell(row=sheet_row, column=6)

    FSA_name = FSA_cell.value

    data[FSA_name] = FSA(FSA_name)

    # Checks if data value is suppressed to prevent assignment errors. If it is suppressed, the
    # value will remain at the default value of -1. This can be easily detected and removed
    # from data comparisons later.
    if cases_per_100_cell.value != '*':
        data[FSA_name].total_covid_cases_per_100 = cases_per_100_cell.value

    if hospitalizations_per_1000_cell.value != '*':
        data[FSA_name].total_covid_hospitalizations_per_1000 = hospitalizations_per_1000_cell.value

    if deaths_per_1000_cell.value != '*':
        data[FSA_name].total_covid_deaths_per_1000 = deaths_per_1000_cell.value

    data[FSA_name].percent_1_dose = percent_1_dose_cell.value

    sheet_row += 1

# Data now contains:
# Cases
# Hospitalizations
# Deaths
# 1-dose vaccinations
# for all Toronto postal codes! Awesome!

# Here, we add the 2-dose vaccination rate
ICES_2_dose_sheet_obj = ICES_wb_obj['2 doses by FSA']

sheet_row = 277
while sheet_row <= 372:
    FSA_cell = ICES_2_dose_sheet_obj.cell(row=sheet_row, column=1)
    percent_2_doses_cell = ICES_2_dose_sheet_obj.cell(row=sheet_row, column=3)

    data[FSA_cell.value].percent_2_doses = percent_2_doses_cell.value

    sheet_row += 1


# Alright! Data now contains:
# Cases
# Hospitalizations
# Deaths
# 1-dose vaccinations
# 2-dose vaccinations
# for all Toronto postal codes! Now, we need population, number of restaurants, and infractions.

# Now, we move to population!
# I have absolutely no idea why the file is called that. It just is.
with open('T120120211212055123.csv', newline='') as population_csv:
    popreader = csv.reader(population_csv, delimiter=',')
    for row in popreader:
        if len(row) == 0:  # This only happens at the end of the file, and stops it from complaining
            break
        if row[0] in data:
            data[row[0]].population = int(row[4])

# All right! Data now contains:
# Cases
# Hospitalizations
# Deaths
# 1-dose vaccinations
# 2-dose vaccinations
# Population
# For all Toronto postal codes! All that's left is number of restaurants and infractions.

# This next bit is very tricky. To add the infractions, we need to know what FSA each restaurant
# is in. But the dinesafe dataset provides the restaurant's street address,
# latitude and longitude, but NOT postal code. This is very annoying since the edges of postal
# codes are very jagged and there is no good way to describe the boundaries. How do we proceed?
# We use the geopy module to get the Nominatim geolocator to ping various map APIs with the
# coordinates to each restaurant! It will send us back a variety of information about those
# coordinates, including their postal code! We can then substring this for our FSA, and then
# map infractions to it and increase the number of restaurants contained within.
# ==================================================================================================
#                                           HOWEVER
# ==================================================================================================
# The Nominatim geolocator API (since I do not have the hardware capacity to download the whole
# thing) has a terms of use that specifies an absolute maximum of one request per second. A quick
# back-of-the-envelope calculation tells us that there are over 17000 restaurants and food-serving
# establishments in the Dinesafe dataset, meaning that at 1 second between requests,
# it will take almost 5 AND A HALF HOURS to do the entire thing. This is why I am
# using this Python file to prepare the data ahead of time into a nice, pre-made file for graphing.
# While this code is a large portion of my project, unless you have 5 hours to spare and a stable
# internet connection, DO NOT RUN THIS FILE.


# Opening the ds.xml data set as a string
dinesafe_string: str
with open('ds.xml', 'r') as file:
    dinesafe_string = file.read().rstrip()

# Strips weird french characters that were breaking the xml reader
cleaned_dinesafe_string = unidecode.unidecode(dinesafe_string)

# open Dinesafe data with minidom
dinesafe_data = minidom.parseString(cleaned_dinesafe_string)
# Create a list of all establishments in the Dinesafe data
establishment_list = dinesafe_data.getElementsByTagName('ESTABLISHMENT')
for establishment in establishment_list:
    # Get latitude and longitude as floats.
    latitude = float(establishment.getElementsByTagName('LATITUDE')[0].childNodes[0].data)
    longitude = float(establishment.getElementsByTagName('LONGITUDE')[0].childNodes[0].data)
    # Get list of infractions by this particular establishment.
    infraction_list = establishment.getElementsByTagName('SEVERITY')

    # Tracking infractions per restaurant
    minor_infractions = 0
    significant_infractions = 0
    crucial_infractions = 0

    # Seeing how many of each kind of infraction the restaurant has
    for infraction in infraction_list:
        if str(infraction.childNodes[0].data) == 'M - Minor':
            minor_infractions += 1
        elif str(infraction.childNodes[0].data) == 'S - Significant':
            significant_infractions += 1
        elif str(infraction.childNodes[0].data) == 'C - Crucial':
            crucial_infractions += 1

    # Formatting coordinates to feed them into pings to Nominatim
    formatted_coordinates = str(str(latitude) + ', ' + str(longitude))

    # Pinging Nominatim geolocator API. This is the part that makes the loop take 5 hours.
    # Setting up user token to give Nominatim
    geolocator = Nominatim(user_agent="restaurant-covid-analysis")
    # Telling Nominatim I am searching for address/postal code by coordinates, not vice versa.
    # Also, setting up RateLimiter to prevent >1 ping per second
    reverse = RateLimiter(geolocator.reverse, min_delay_seconds=1, max_retries=5)
    # Call to Nominatim API
    location = reverse(formatted_coordinates)
    # Checking to make sure location is recieved properly, and that it has a valid postal code in
    # The Toronto area. If these conditions are not met, the result is discarded. In the end, only
    # A VERY small fraction of cases get discarded.
    if location is not None and 'address' in location.raw and 'postcode' in location.raw['address']:
        FSA = str.split(location.raw['address']['postcode'])[0]
        if FSA in data:
            data[FSA].number_of_minor_infractions += minor_infractions
            data[FSA].number_of_significant_infractions += significant_infractions
            data[FSA].number_of_crucial_infractions += crucial_infractions
            data[FSA].number_of_restaurants += 1


# Fantastic! All the data is now properly recorded inside the data dict!

# Let's write it out to a .csv file, so we can access it later!
with open('cleaned_data.csv', 'w', newline='') as csvfile:
    my_writer = csv.writer(csvfile)
    my_writer.writerow(['FSA', 'Number of Restaurants', 'Number of Minor Infractions',
                        'Number of Significant Infractions', 'Number of Crucial Infractions',
                        'Total COVID-19 Cases per 100 people',
                        'Total COVID-19 Hospitalizations per 1000 people',
                        'Total COVID-19 Deaths per 1000 people',
                        'Population', 'Percent With At Least 1 Dose', 'Percent With 2 Doses'])
    for FSA in data:  # Write all the data for this FSA into the csv cleaned_data file
        # Variable assignment for cleanliness
        name = data[FSA].name
        number_of_restaurants = data[FSA].number_of_restaurants
        number_of_minor_infractions = data[FSA].number_of_minor_infractions
        number_of_significant_infractions = data[FSA].number_of_significant_infractions
        number_of_crucial_infractions = data[FSA].number_of_crucial_infractions
        total_covid_cases_per_100 = data[FSA].total_covid_cases_per_100
        total_covid_hospitalizations_per_1000 = data[FSA].total_covid_hospitalizations_per_1000
        total_covid_deaths_per_1000 = data[FSA].total_covid_deaths_per_1000
        population = data[FSA].population
        percent_1_dose = data[FSA].percent_1_dose
        percent_2_doses = data[FSA].percent_2_doses

        # Actually write the row
        my_writer.writerow([name, number_of_restaurants, number_of_minor_infractions,
                            number_of_significant_infractions, number_of_crucial_infractions,
                            total_covid_cases_per_100, total_covid_hospitalizations_per_1000,
                            total_covid_deaths_per_1000, population, percent_1_dose,
                            percent_2_doses])
